import openpyxl,sys,pyperclip
from tkinter import*
import tkinter.messagebox

def ShowSheets(WorkBook):
    SheetNames = WorkBook.get_sheet_names()
    print("There are {} Sheets in the WorkBook".format(len(SheetNames)))
    print("They are:\n\n")
    for i in range(0,len(SheetNames)):
        print(SheetNames[i])
    ActiveSheet = WorkBook.get_active_sheet()
    print(ActiveSheet)
    print("\n\n")

def colToExcel(col): # col is 1 based
    excelCol = str()
    div = col 
    while div:
        (div, mod) = divmod(div-1, 26) # will return (x, 0 .. 25)
        excelCol = chr(mod + 65) + excelCol
    return excelCol

root = Tk()
root.iconbitmap(r'C:\Users\HARI\AppData\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
root.wm_title("Hari Excell Manager")
tkinter.messagebox.showinfo("Hari Technologies","Welcome to HARI EXCEL MANAGER.VIEWER\n->User-Friendly\n\nThanks For Chosing us!!")
root.destroy()
root.mainloop()
print("Enter Addres of the Excel File to Handle with.")
ClipBoard = pyperclip.paste()
if(ClipBoard != "" and ((".xlsx" in ClipBoard )or( ".xlx" in ClipBoard))):
    """print("\n******* INFORMATION *****")
    print(ClipBoard)
    print("We have Detected Text in your Clip Board\nDo you Want to use it?\nPress Enter to Use it :: Press any key and then Enter")
    Temp = input()"""
    obj = Tk()
    obj.iconbitmap(r'C:\Users\HARI\AppDAta\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
    obj.wm_title("Hari Excell Manager")
    Answer = tkinter.messagebox.askquestion("Hari Technologies","We have Detected Text in your Clip Board\n"+ClipBoard+"\nDo you Want to use it?")
    obj.destroy()
    obj.mainloop()
    if(Answer == "yes"):
        Address = ClipBoard
    if(Answer == "no"):
        Address = input()
else:
    Address=input()
root.mainloop()
try:
    WorkBook = openpyxl.load_workbook(Address)
except FileNotFoundError:
    root = Tk()
    root.iconbitmap(r'C:\Users\HARI\AppData\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
    root.wm_title("Hari Excell Manager")
    tkinter.messagebox.showinfo("Hari Technologies","Invalid Address\n\nClick OK to Abort Program")
    root.destroy()
    root.mainloop()
    sys.exit(0)
    
except openpyxl.utils.exceptions.InvalidFileException:
    root = Tk()
    root.iconbitmap(r'C:\Users\HARI\AppData\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
    root.wm_title("Hari Excell Manager")
    tkinter.messagebox.showinfo("Hari Technologies","Invalid Format of File \nOnly Supports .xlx OR .xlsx\n\nClick OK to Abort Program")
    root.destroy()
    root.mainloop()
    sys.exit(0)

ShowSheets(WorkBook)
print("Please Enter a Sheet to Work On..")
print("Press Enter to proceed with Active Sheet")
SheetInput = input()
if(SheetInput != ""):
    Sheet = WorkBook.get_sheet_by_name(SheetInput)
else:
    Sheet = WorkBook.active
    
MaxRow = Sheet.max_row
MaxCol = Sheet.max_column
print("Maximum Number of Rows = {}".format(str(MaxRow)))
print("Maximum Column Name ={}".format(colToExcel(MaxCol)))
print("\n\n")

#obj = Tk()
#obj.iconbitmap(r'C:\Users\HARI\AppDAta\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
#obj.wm_title("Hari Excell Manager")
#Answer_1  = tkinter.messagebox.askquestion("Excel to TEXT Converter","Want To create an Equivalent Text File\n\n(FOR THIS SHEET) ?")
#obj.destroy()
#obj.mainloop()
#if(Answer_1=="no"):#painavi pedithe kinda vunna print("***") varuku indentation choodali.
for Row in range(1,MaxRow+1):
    print("ROW = {}\n\n".format(str(Row)))
    for Col in range(1,MaxCol+1):
        CellName = colToExcel(Col) + str(Row)
        CellValue = Sheet[CellName].value
        if(CellValue != None):
            print("\t"+str(CellName)+"\t\t"+str(CellValue))
            print("\n\n\n")
    print("**************************************** ==== ********************************************************************************************")

root = Tk()
root.iconbitmap(r'C:\Users\HARI\AppData\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
root.wm_title("Hari Excell Manager")
tkinter.messagebox.showinfo("Hari Technologies","Thanks For Chosing us!!")
root.destroy()
root.mainloop()
