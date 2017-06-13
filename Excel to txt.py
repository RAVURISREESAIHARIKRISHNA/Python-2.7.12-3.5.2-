##################    Excel Sheet to TXT Converting ##############################

## NOTE : .txt File Support Only UTF-8 Encoding
## They Do not Support Special Symbols Like Sigma,Pi,THeta and all other like Symblos

# @Copyright "Hari Technologies"

import openpyxl,sys,os,pyperclip
from tkinter import*
import tkinter.messagebox


def colToExcel(col): # col is 1 based
    excelCol = str()
    div = col 
    while div:
        (div, mod) = divmod(div-1, 26) # will return (x, 0 .. 25)
        excelCol = chr(mod + 65) + excelCol
    return excelCol



root = Tk()
root.iconbitmap(r'C:\Users\HARI\AppData\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
root.wm_title("Hari Excell Manager")
tkinter.messagebox.showinfo("Hari Technologies","Welcome to HARI EXCEL to TXT COnverter\n->User-Friendly\n\nThanks For Chosing us!!")
root.destroy()
root.mainloop()


print("Enter Address")




ClipBoard = pyperclip.paste()
if(ClipBoard != "" and ((".xlsx" in ClipBoard )or( ".xlx" in ClipBoard))):
    """print("\n******* INFORMATION *****")
    print(ClipBoard)
    print("We have Detected Text in your Clip Board\nDo you Want to use it?\nPress Enter to Use it :: Press any key and then Enter")
    Temp = input()"""
    obj = Tk()
    obj.iconbitmap(r'C:\Users\HARI\AppDAta\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
    obj.wm_title("Hari Excell Manager")
    Answer = tkinter.messagebox.askquestion("Hari Technologies","We have Detected Text in your Clip Board\n"+ClipBoard+"\nDo you Want to use it?")
    obj.destroy()
    obj.mainloop()
    if(Answer == "yes"):
        Address = ClipBoard
    if(Answer == "no"):
        Address = input()
else:
    Address=input()
root.mainloop()
#Address = input()


try:
    WorkBook = openpyxl.load_workbook(Address)
except FileNotFoundError:
    root = Tk()
    root.iconbitmap(r'C:\Users\HARI\AppData\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
    root.wm_title("Hari Excell Manager")
    tkinter.messagebox.showinfo("Hari Technologies","Invalid Address\n\nClick OK to Abort Program")
    root.destroy()
    root.mainloop()
    sys.exit(0)
    
except openpyxl.utils.exceptions.InvalidFileException:
    root = Tk()
    root.iconbitmap(r'C:\Users\HARI\AppData\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
    root.wm_title("Hari Excell Manager")
    tkinter.messagebox.showinfo("Hari Technologies","Invalid Format of File \nOnly Supports .xlx OR .xlsx\n\nClick OK to Abort Program")
    root.destroy()
    root.mainloop()
    sys.exit(0)

#WorkBook = openpyxl.load_workbook(Address)

SheetNames = WorkBook.get_sheet_names()
print("There are {} Sheets".format(len(SheetNames)))
print("They are")
for i in range(0,len(SheetNames)):
    print(SheetNames[i])
ActiveSheet = WorkBook.active
print("Active Sheet:")
print(ActiveSheet)


print("Please Enter Sheet Name")
SelectedSheetName = input()
if(SelectedSheetName not in SheetNames):
    # Abort Code
    root = Tk()
    root.iconbitmap(r'C:\Users\HARI\AppData\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
    root.wm_title("Hari Excell Manager")
    tkinter.messagebox.showinfo("Hari Excel to TXT Converter","Sheet Name not Found \n\nClick OK to Abort")
    root.destroy()
    root.mainloop()

    #print("Name Not Found")
    sys.exit(0)

Sheet = WorkBook.get_sheet_by_name(SelectedSheetName)

MaxRow = Sheet.max_row
MaxCol = Sheet.max_column

obj = Tk()
obj.iconbitmap(r'C:\Users\HARI\AppDAta\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
obj.wm_title("Hari Excell Manager")
Answer = tkinter.messagebox.askquestion("Excel to TEXT Converter","Do you Want to Save the txt File in Same Path OR enter New Address ?")
obj.destroy()
obj.mainloop()

#print("Do you Want to Save the txt File in Same Path OR enter New Address ?")
#Answer = input()
if(Answer == "yes"):
    TempAddress = os.path.dirname(Address)
    FileAddress = TempAddress + '//'+SelectedSheetName + '.txt'

if(Answer == "no"):
    FileAddress = input()
File = open(FileAddress , "w")
File.write('@ Copyright \"Hari Technologies\"\nExcel to TXT Converter')
File.write('**********************************"+SelectedSheetName+"************************************\n\n\n')

for Row in range(1,MaxRow+1):
    File.write("Row "+str(Row)+"\n")
    for Col in range(1,MaxCol):
        CellName = colToExcel(Col) + str(Row)
        CellValue = str(Sheet[CellName].value)
        DataString = CellName + "    " + CellValue
        if(Sheet[CellName].value != None):
            try:
                File.write(DataString)
            except UnicodeEncodeError:
                root = Tk()
                root.iconbitmap(r'C:\Users\HARI\AppData\Local\Programs\Python\Python35-32\DLLs\My_Icon.ico')
                root.wm_title("Hari Excell Manager")
                tkinter.messagebox.showinfo("Hari Excel to TXT Converter","This Sheet has Some Character which Does not Support UTF-8 Encoding.\nEG:Sigma,Pi...\n\nClick OK to Abort")
                root.destroy()
                root.mainloop()
                
        






        File.write("\n")
        File.write("\n")
    File.write("<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>\n\n\n")
File.close()
