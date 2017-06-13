import openpyxl,sys

def ShowSheets(WorkBook):
    SheetNames = WorkBook.get_sheet_names()
    print("There are {} Sheets".format(len(SheetNames)))
    print("There are")
    for i in range(0,len(SheetNames)-1):
        print(SheetNames[i])
    print("The Active Sheet is")
    ActiveSheet = WorkBook.get_active_sheet()
    print(ActiveSheet)

"""def MaxRow(Sheet):
    return(Sheet.max_row)

def MaxCol(Sheet):
    return(Sheet.max_column)"""

def colToExcel(col): # col is 1 based
    excelCol = str()
    div = col 
    while div:
        (div, mod) = divmod(div-1, 26) # will return (x, 0 .. 25)
        excelCol = chr(mod + 65) + excelCol
    return excelCol
    
"""def SheetDisplay(Sheet):
    SheetName = WorkBook.get_sheet_by_name(Sheet)
    
    RowSize = SheetName.max_row
    ColSize = SheetName.max_column
    for Row in range(1,RowSize):
        for Col in range(1,ColSize):
            CellName = colToExcel(Col)+str(Row)
            print(str(CellName)+"\t"+str(Sheet[str(CellName)].value))
        print("*****************  =====  ***************")
        """"
    

def SelectSheet(WorkBook):
    ShowSheets(WorkBook)
    print("Please Enter Sheet Name to Select\nOR\nPress Enter to Procced with Active Sheet")
    SheetName = input()
    if(SheetName == ""):
        ActiveSheet = WorkBook.get_active_sheet()
        return ActiveSheet
    else:
        SheetNames = WorkBook.get_sheet_names()
        if(SheetName in SheetNames):
            return SheetName
        else:
            print("Process Aborted,Please Restart the PRogram")
            sys.exit(0)
            #return SheetName
    
print("Please Enter a vald Address of an Excell File")
Address = input()
WorkBook = openpyxl.load_workbook(Address)
while(True):
    print("**** MENU *****")
    print("1. Show Sheets")
    print("2. Select Sheet")
    #print("3. Display Sheet")
    print("Enter Choice")
    ch = int(input())
    if(ch not in [1,2]):
        sys.exit(0)
    if(ch == 1):
        ShowSheets(WorkBook)
    if(ch == 2):
        Sheet1 = SelectSheet(WorkBook)
    
    print("Selected Sheeet is :"+Sheet1)
        
